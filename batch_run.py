"""Orchestrazione automatica per lotti numerosi.

Divide i PDF in piccoli gruppi, richiama run.py, applica pause e retry a livello
di batch e consolida gli Excel intermedi in un unico file finale.
"""

import argparse
import shutil
import subprocess
import sys
import time
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


def run_batch(
    batch_folder: Path,
    model: str,
    output_file: Path,
    retries: int,
    retry_wait: int,
) -> None:
    """Esegue run.py su un gruppo di PDF, riprovando in caso di errore."""

    command = [
        sys.executable,
        "run.py",
        str(batch_folder),
        "--model",
        model,
        "--out",
        str(output_file),
    ]

    for attempt in range(1, retries + 1):
        print(f"\nTentativo {attempt}/{retries} per {batch_folder.name}")
        result = subprocess.run(command, check=False)

        if result.returncode == 0 and output_file.exists():
            print(f"Gruppo completato: {batch_folder.name}")
            return

        if attempt < retries:
            print(
                f"Errore nel gruppo {batch_folder.name}. "
                f"Nuovo tentativo tra {retry_wait} secondi..."
            )
            time.sleep(retry_wait)

    raise RuntimeError(
        f"Il gruppo {batch_folder.name} non è stato completato "
        f"dopo {retries} tentativi."
    )


def copy_cell(source_cell, destination_cell) -> None:
    """Copia valore e formattazione di una cella Excel."""

    destination_cell.value = source_cell.value

    if source_cell.has_style:
        destination_cell.font = copy(source_cell.font)
        destination_cell.fill = copy(source_cell.fill)
        destination_cell.border = copy(source_cell.border)
        destination_cell.alignment = copy(source_cell.alignment)
        destination_cell.number_format = source_cell.number_format
        destination_cell.protection = copy(source_cell.protection)


def merge_excel_files(files: list[Path], final_output: Path) -> None:
    """Unisce i file Excel dei singoli gruppi in un unico file."""

    if not files:
        raise RuntimeError("Non ci sono file Excel da unire.")

    final_output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(files[0], final_output)

    destination_workbook = load_workbook(final_output)
    destination_sheet = destination_workbook.active

    for source_file in files[1:]:
        source_workbook = load_workbook(source_file)
        source_sheet = source_workbook.active

        # La riga 1 contiene le intestazioni, quindi si parte dalla riga 2.
        for source_row in source_sheet.iter_rows(min_row=2):
            destination_row_number = destination_sheet.max_row + 1

            for column_number, source_cell in enumerate(source_row, start=1):
                destination_cell = destination_sheet.cell(
                    row=destination_row_number,
                    column=column_number,
                )
                copy_cell(source_cell, destination_cell)

        source_workbook.close()

    destination_workbook.save(final_output)
    destination_workbook.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Elabora automaticamente i PDF in piccoli gruppi, "
            "gestisce le pause e crea un unico Excel."
        )
    )
    parser.add_argument("folder", help="Cartella contenente tutti i PDF")
    parser.add_argument("--model", required=True, help="Modello Gemini da utilizzare")
    parser.add_argument(
        "--out",
        default="output/Dati_Pack.xlsx",
        help="File Excel finale",
    )
    parser.add_argument(
        "--batch-size",
        type=int,
        default=5,
        help="Numero di PDF per gruppo",
    )
    parser.add_argument(
        "--pause",
        type=int,
        default=60,
        help="Secondi di pausa tra i gruppi",
    )
    parser.add_argument(
        "--retries",
        type=int,
        default=4,
        help="Numero massimo di tentativi per ogni gruppo",
    )
    parser.add_argument(
        "--retry-wait",
        type=int,
        default=120,
        help="Secondi di attesa dopo un errore",
    )
    args = parser.parse_args()

    source_folder = Path(args.folder).resolve()
    final_output = Path(args.out).resolve()

    if not source_folder.exists():
        raise FileNotFoundError(f"La cartella non esiste: {source_folder}")

    pdf_files = sorted(
        file
        for file in source_folder.iterdir()
        if file.is_file() and file.suffix.lower() == ".pdf"
    )

    if not pdf_files:
        raise RuntimeError(f"Nessun PDF trovato nella cartella: {source_folder}")

    if args.batch_size < 1:
        raise ValueError("--batch-size deve essere almeno 1.")

    work_folder = Path("_automatic_batches").resolve()
    batch_output_folder = Path("output/_batch_files").resolve()

    # Elimina eventuali residui di un'esecuzione precedente.
    if work_folder.exists():
        shutil.rmtree(work_folder)
    if batch_output_folder.exists():
        shutil.rmtree(batch_output_folder)

    work_folder.mkdir(parents=True)
    batch_output_folder.mkdir(parents=True)

    total_batches = (len(pdf_files) + args.batch_size - 1) // args.batch_size

    print(f"PDF trovati: {len(pdf_files)}")
    print(f"Dimensione gruppi: {args.batch_size}")
    print(f"Gruppi totali: {total_batches}")

    generated_files: list[Path] = []

    try:
        for batch_index in range(total_batches):
            start = batch_index * args.batch_size
            end = start + args.batch_size
            current_files = pdf_files[start:end]

            batch_number = batch_index + 1
            batch_folder = work_folder / f"batch_{batch_number:03d}"
            batch_folder.mkdir()

            print(
                f"\nPreparazione gruppo {batch_number}/{total_batches}: "
                f"{len(current_files)} PDF"
            )

            for pdf_file in current_files:
                shutil.copy2(pdf_file, batch_folder / pdf_file.name)

            batch_output = (
                batch_output_folder
                / f"Dati_Pack_batch_{batch_number:03d}.xlsx"
            )

            run_batch(
                batch_folder=batch_folder,
                model=args.model,
                output_file=batch_output,
                retries=args.retries,
                retry_wait=args.retry_wait,
            )
            generated_files.append(batch_output)

            if batch_number < total_batches:
                print(
                    f"Pausa automatica di {args.pause} secondi "
                    "prima del gruppo successivo..."
                )
                time.sleep(args.pause)

        print("\nUnione automatica dei file Excel...")
        merge_excel_files(generated_files, final_output)

        print("\nElaborazione completata.")
        print(f"File finale: {final_output}")

    finally:
        # Elimina solo le cartelle temporanee contenenti le copie dei PDF.
        if work_folder.exists():
            shutil.rmtree(work_folder)


if __name__ == "__main__":
    main()
