"""
Pipeline di estrazione dati da fustelle pack -> riga strutturata.

Architettura ibrida:
  ANCORE DETERMINISTICHE (costo 0, affidabilita' ~100%)
    - nome file : EAN-13 (con check-digit) + dimensioni scatola + nome/variante
    - testo vivo: unica stringa non vettorializzata ("FR 21 7 PAP CPE") -> PAP21/CPE07
    - costanti brand: fabbricante/importatore (uguali per ogni prodotto MySecretCase)
  VISION LLM (gemini_vision.py)
    - tutto il resto (LOT, batteria, materiale, IPX, dimensioni prodotto, feature, icone)
  GUARD DI VALIDAZIONE
    - EAN check-digit + cross-check nome-file vs cifre stampate
    - accordo doppia-lettura vision
    - regole di dominio (es. impermeabilita' senza IPX su prodotto elettronico)
    - feature senza colonna -> flag; campo illeggibile -> NEEDS_REVIEW
"""
from __future__ import annotations
import io, os, re, warnings
import pdfplumber
from PIL import Image
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

# ---- 34 colonne della mappatura di riferimento (ordine A..AH) -----------------
COLS = [
    "Nome del fabbricante","Indirizzo del fabbricante","Nome dell'importatore",
    "Indirizzo dell'importatore","Tipo o modello","Numero di serie / lotto","Lotto",
    "Simbolo CE","Simbolo RAEE","Simbolo UKCA","Simbolo TRIMAN",
    "Simbolo smaltimento spagnolo","Simboli materiali smaltimento","QR code Junker",
    "Simbolo garanzia 2 anni","Simbolo libretto informativo",
    "Capacità batteria e tensione nominale","Impermeabilità","Materiale",
    "Modalità di ricarica","Dimensioni","N. vibrazioni","N. velocità",
    "N. modalità suzione","N. modalità tapping","N. modalità rotazione",
    "Strap-on compatibile","Funzione riscaldante","Codice ASIN",
    "Codice smaltimento scatola","Codice smaltimento sacchetto",
    "Codice smaltimento doypack","Contenuto TRIMAN corretto","Sexy Ideas",
]
SRC = ["C","C","C","C","C","V","C","V","V","V","V","V","L","V","V","V",
       "V","V","V","V","V","V","V","V","V","V","V","V","X","L","L","L","V","V"]

CONST = {
    "Nome del fabbricante":"MySecretCase s.r.l.",
    "Indirizzo del fabbricante":"Corso C. Colombo 7 - Milano 20144",
    "Nome dell'importatore":"MySecretCase s.r.l.",
    "Indirizzo dell'importatore":"Corso C. Colombo 7 - Milano 20144",
    "Tipo o modello":"N/A", "Lotto":"N/A",
}
KNOWN_FEATURES = {"vibrazioni","velocità","velocita","suzione","tapping","movimento",
                  "rotazione","strap-on","strap","riscaldante","impermeabile","waterproof",
                  "ricarica","charge","materiale","silicone","garanzia","warranty",
                  # ignorate per scelta del cliente (nessuna colonna, nessun flag):
                  "telecomandato","telecomandata","telecomando","remote","remote control"}

def tick(b: bool) -> str: return "✅" if b else "❌"
def txt(s: str) -> str:   return s.strip() if s and s.strip() else "❌"

# ---------- 1. NOME FILE (deterministico) --------------------------------------
FNAME = re.compile(r'^(\d{13})_(\d+)x(\d+)x(\d+)_(.+?)\.(?:pdf|ai)$', re.I)

def validate_ean13(ean: str) -> bool:
    if not (len(ean) == 13 and ean.isdigit()): return False
    d = [int(x) for x in ean]
    chk = (10 - (sum(d[i]*(1 if i % 2 == 0 else 3) for i in range(12)) % 10)) % 10
    return chk == d[12]

def parse_filename(name: str) -> dict | None:
    m = FNAME.match(os.path.basename(name))
    if not m: return None
    ean, L, W, H, prod = m.groups()
    prod = prod.replace("_", " ").replace(" - ", " — ").strip()
    return {"ean": ean, "ean_valid": validate_ean13(ean),
            "box_mm": f"{L}x{W}x{H}", "product": prod}

# ---------- 2. TESTO VIVO riciclo (deterministico, spaziale) -------------------
def extract_recycling_codes(pdf_path: str) -> dict:
    with pdfplumber.open(pdf_path) as pdf:
        words = pdf.pages[0].extract_words()
    labels = [w for w in words if w["text"] in ("PAP", "CPE")]
    nums   = [w for w in words if re.fullmatch(r"\d+", w["text"])]
    out = {}
    for lb in labels:
        cx = (lb["x0"] + lb["x1"]) / 2
        cand = min(nums, key=lambda n: abs((n["x0"]+n["x1"])/2 - cx)
                                       + (0 if n["top"] < lb["top"] else 999), default=None)
        if cand: out[lb["text"]] = cand["text"].zfill(2)
    return {"scatola": f"PAP{out['PAP']}" if "PAP" in out else "",
            "sacchetto": f"CPE{out['CPE']}" if "CPE" in out else ""}

# ---------- 3. RASTER pagina 1 -------------------------------------------------
def rasterize_page1(pdf_path: str, dpi: int = 200) -> bytes:
    import subprocess, tempfile, glob
    with tempfile.TemporaryDirectory() as d:
        subprocess.run(["pdftoppm","-jpeg","-r",str(dpi),"-f","1","-l","1",
                        pdf_path, os.path.join(d,"p")], check=True,
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        f = sorted(glob.glob(os.path.join(d,"p*.jpg")))[0]
        return open(f,"rb").read()

# ---------- 4. ASSEMBLAGGIO riga + GUARD --------------------------------------
def assemble_row(fn: dict, rec: dict, vis: dict, disagree: list[str]) -> dict:
    r = dict(CONST)
    r["Numero di serie / lotto"] = txt(vis.get("lot",""))
    r["Simbolo CE"]              = tick(vis.get("ce"))
    r["Simbolo RAEE"]            = tick(vis.get("raee"))
    r["Simbolo UKCA"]            = tick(vis.get("ukca"))
    r["Simbolo TRIMAN"]          = tick(vis.get("triman"))
    r["Simbolo smaltimento spagnolo"] = tick(vis.get("spanish_disposal"))
    r["Simboli materiali smaltimento"] = " / ".join(x for x in
        (rec["scatola"], rec["sacchetto"]) if x) or "❌"
    r["QR code Junker"]           = tick(vis.get("junker_qr"))
    r["Simbolo garanzia 2 anni"]  = tick(vis.get("warranty_2y"))
    r["Simbolo libretto informativo"] = tick(vis.get("booklet"))
    r["Capacità batteria e tensione nominale"] = txt(vis.get("battery",""))
    r["Impermeabilità"]  = txt(vis.get("waterproof",""))
    r["Materiale"]       = txt(vis.get("material",""))
    r["Modalità di ricarica"] = txt(vis.get("charge_mode",""))
    r["Dimensioni"]      = txt(vis.get("product_dimensions",""))
    r["N. vibrazioni"]   = txt(vis.get("vibration",""))
    r["N. velocità"]     = txt(vis.get("speed",""))
    r["N. modalità suzione"]   = txt(vis.get("suction",""))
    r["N. modalità tapping"]   = txt(vis.get("tapping",""))
    r["N. modalità rotazione"] = txt(vis.get("rotation",""))
    r["Strap-on compatibile"]  = tick(vis.get("strap_on"))
    r["Funzione riscaldante"]  = tick(vis.get("heating"))
    r["Codice ASIN"]           = "N/D"          # non presente sul pack (vedi README §11)
    r["Codice smaltimento scatola"]   = rec["scatola"] or "❌"
    r["Codice smaltimento sacchetto"] = rec["sacchetto"] or "❌"
    r["Codice smaltimento doypack"]   = "❌"
    r["Contenuto TRIMAN corretto"]    = txt(vis.get("triman_content",""))
    r["Sexy Ideas"] = "✅" if (vis.get("sexy_ideas") or "").strip() else "❌"

    # ---- GUARD ----
    flags, conf = [], "Alta"
    if not fn["ean_valid"]:
        flags.append("EAN-13 check-digit NON valido"); conf = "Bassa"
    if r["Numero di serie / lotto"] == "❌":
        flags.append("LOT illeggibile -> NEEDS_REVIEW"); conf = "Bassa"
    if r["Dimensioni"] == "❌":
        flags.append("Dimensioni prodotto assenti -> NEEDS_REVIEW"); conf = "Bassa"
    if disagree:
        flags.append("Vision incoerente su: " + ", ".join(disagree)); conf = "Media"
    # regola di dominio: impermeabilita' senza IPX su prodotto elettronico
    electronic = r["Capacità batteria e tensione nominale"] != "❌"
    wp = r["Impermeabilità"]
    if wp not in ("❌",) and "IPX" not in wp.upper():
        if electronic:
            flags.append("Impermeabilità senza codice IPX (prodotto elettronico)"); conf = "Media"
        else:
            flags.append("Impermeabilità senza codice IPX; non elettronico -> Q/T/V ❌ attesi")
            conf = "Media" if conf == "Alta" else conf
    # feature con icona ma senza colonna nel mapping
    for feat in (vis.get("other_features_seen","") or "").split(","):
        f = feat.strip()
        if f and f.lower() not in KNOWN_FEATURES:
            flags.append(f"Feature '{f}' presente ma senza colonna dedicata")

    r["_EAN"] = fn["ean"]; r["_BOX"] = fn["box_mm"]
    r["_CONF"] = conf; r["_FLAG"] = "; ".join(flags) if flags else "—"
    r["_PROD"] = fn["product"]
    return r

# ---------- 5. ORCHESTRAZIONE cartella ----------------------------------------
def process_folder(folder: str, extractor) -> list[dict]:
    rows = []
    for name in sorted(os.listdir(folder)):
        if not name.lower().endswith((".pdf", ".ai")): continue
        path = os.path.join(folder, name)
        fn = parse_filename(name)
        if fn is None:
            rows.append({"_EAN":"", "_PROD":name, "_CONF":"Bassa",
                         "_FLAG":"Nome file non conforme al pattern EAN_LxWxH_Nome"})
            continue
        rec = extract_recycling_codes(path)
        img = rasterize_page1(path)
        vis, disagree = extractor.extract(img, {"ean": fn["ean"], "product": fn["product"]})
        rows.append(assemble_row(fn, rec, vis, disagree))
    return rows

# ---------- 6. SINK: Google Sheet (xlsx compatibile) --------------------------
def write_xlsx(rows: list[dict], out_path: str):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Dati"
    A = "Arial"
    GREEN = PatternFill("solid", fgColor="C6EFCE")
    BLUE  = PatternFill("solid", fgColor="BDD7EE")
    RED   = PatternFill("solid", fgColor="FFC7CE")
    GREY  = PatternFill("solid", fgColor="E7E6E6")
    fill_for = {"F":GREEN,"L":GREEN,"C":GREEN,"V":BLUE,"X":RED}
    thin = Side(style="thin", color="BFBFBF"); bd = Border(thin,thin,thin,thin)

    # 34 colonne ufficiali (identiche all'esempio) + blocco note di servizio (grigio) a destra.
    # L'EAN NON e' un campo ufficiale: resta come riferimento, chiaramente separato.
    HELP = ["EAN (rif.)","Dimensioni scatola (rif.)","Confidenza","Flag / Note","Prodotto (rif.)"]
    header  = COLS + HELP
    srcline = SRC + [""]*len(HELP)            # helper -> nessun colore fonte (grigio)
    n_spec = len(COLS)
    for j,(h,s) in enumerate(zip(header,srcline),1):
        c = ws.cell(1,j,h); c.font = Font(name=A,bold=True,size=10)
        c.alignment = Alignment(wrap_text=True,vertical="center"); c.border = bd
        c.fill = fill_for.get(s,GREY)
    for i,d in enumerate(rows,2):
        vals = [d.get(c,"") for c in COLS] + \
               [d.get("_EAN",""), d.get("_BOX",""), d.get("_CONF",""), d.get("_FLAG",""), d.get("_PROD","")]
        for j,v in enumerate(vals,1):
            c = ws.cell(i,j,v); c.border = bd
            c.font = Font(name=A,size=10)
            c.alignment = Alignment(vertical="center", wrap_text=(j>n_spec))
    for j in range(1,len(header)+1):
        ws.column_dimensions[get_column_letter(j)].width = min(max(len(header[j-1])+2,12),30)
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 42

    # foglio provenienza
    lg = wb.create_sheet("Legenda_Provenienza")
    def put(r,c,v,b=False,f=None):
        x=lg.cell(r,c,v); x.font=Font(name=A,bold=b,size=10)
        if f:x.fill=f; x.alignment=Alignment(vertical="center",wrap_text=True); return x
    put(1,1,"Legenda & provenienza",True)
    put(3,1,"✅ presente sul pack · ❌ assente / non applicabile")
    put(4,2,"Verde = deterministico (nome file / testo vivo / costante) ~100%",f=GREEN)
    put(5,2,"Blu = letto da LLM vision",f=BLUE)
    put(6,2,"Rosso = non sul pack, richiede fonte esterna",f=RED)
    put(7,2,"Grigio = note di servizio (rif./confidenza/flag), NON tra i campi richiesti",f=GREY)
    put(9,1,"Campo (34 ufficiali, ordine come da esempio)",True); put(9,2,"Fonte",True)
    lab={"F":"Nome file","L":"Testo vivo","V":"Vision LLM","C":"Costante","X":"Esterna"}
    r=10
    for col,s in zip(COLS,SRC): put(r,1,col); put(r,2,lab[s]); r+=1
    for L,w in [("A",34),("B",16)]: lg.column_dimensions[L].width=w
    wb.save(out_path)
